import flet as ft
import sqlite3
import openpyxl
from fpdf import FPDF
from datetime import datetime
import os
import traceback

def main(page: ft.Page):
    try:
        page.title = "Sales Note Manager"
        page.theme_mode = ft.ThemeMode.LIGHT
        page.scroll = "auto"
        
        # --- FIX: Android Writable Database Path ---
        # Android blocks writing to the root folder, so we use the safe HOME folder
        db_dir = os.environ.get("HOME", os.getcwd())
        db_path = os.path.join(db_dir, "sales_database.db")
        
        # --- Database Setup ---
        conn = sqlite3.connect(db_path, check_same_thread=False)
        c = conn.cursor()
        c.execute('CREATE TABLE IF NOT EXISTS companies (name TEXT UNIQUE)')
        c.execute('CREATE TABLE IF NOT EXISTS items (name TEXT UNIQUE)')
        c.execute('''CREATE TABLE IF NOT EXISTS transactions 
                     (id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, company TEXT, item TEXT, 
                     qty REAL, rate REAL, total REAL, gem_order TEXT)''')
        conn.commit()

        def fetch_all(table):
            c.execute(f"SELECT name FROM {table}")
            return [row[0] for row in c.fetchall()]

        # --- UI Elements: Create Note ---
        company_dropdown = ft.Dropdown(label="Select Company")
        item_dropdown = ft.Dropdown(label="Select Item")
        qty_input = ft.TextField(label="Quantity (Allows Negative)", value="1", keyboard_type="number")
        rate_input = ft.TextField(label="Rate per 1 Qty", value="0", keyboard_type="number")
        gem_checkbox = ft.Checkbox(label="Order placed via GeM Portal?")

        def load_dropdowns():
            company_dropdown.options = [ft.dropdown.Option(name) for name in fetch_all("companies")]
            item_dropdown.options = [ft.dropdown.Option(name) for name in fetch_all("items")]
            page.update()

        def show_snack(text):
            page.snack_bar = ft.SnackBar(ft.Text(text))
            page.snack_bar.open = True
            page.update()

        def save_note(e):
            if not company_dropdown.value or not item_dropdown.value:
                show_snack("Please select a Company and an Item!")
                return
            try:
                qty = float(qty_input.value)
                rate = float(rate_input.value)
                total = qty * rate
                gem_status = "Yes" if gem_checkbox.value else "No"
                date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                c.execute('''INSERT INTO transactions (date, company, item, qty, rate, total, gem_order) 
                             VALUES (?, ?, ?, ?, ?, ?, ?)''', 
                          (date_str, company_dropdown.value, item_dropdown.value, qty, rate, total, gem_status))
                conn.commit()
                show_snack(f"Saved successfully! Total: {total}")
            except Exception:
                show_snack("Error: Invalid numbers entered.")

        # --- UI Elements: DB Management ---
        new_company = ft.TextField(label="New Company Name")
        new_item = ft.TextField(label="New Item Name")

        def add_company(e):
            if new_company.value.strip():
                try:
                    c.execute("INSERT INTO companies (name) VALUES (?)", (new_company.value.strip(),))
                    conn.commit()
                    load_dropdowns()
                    new_company.value = ""
                    show_snack("Company added!")
                except:
                    show_snack("Company already exists!")

        def add_item(e):
            if new_item.value.strip():
                try:
                    c.execute("INSERT INTO items (name) VALUES (?)", (new_item.value.strip(),))
                    conn.commit()
                    load_dropdowns()
                    new_item.value = ""
                    show_snack("Item added!")
                except:
                    show_snack("Item already exists!")

        # --- Export / File Pickers ---
        def get_data():
            c.execute("SELECT date, company, item, qty, rate, total, gem_order FROM transactions ORDER BY id DESC")
            return c.fetchall()

        def save_excel(e: ft.FilePickerResultEvent):
            if e.path:
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append(["Date", "Company", "Item", "Quantity", "Rate", "Total", "GeM_Order"])
                    for row in get_data():
                        ws.append(list(row))
                    wb.save(e.path)
                    show_snack("Excel Exported safely!")
                except Exception as ex:
                    show_snack(f"Export Error: {ex}")

        def save_pdf(e: ft.FilePickerResultEvent):
            if e.path:
                try:
                    pdf = FPDF()
                    pdf.add_page()
                    pdf.set_font("Helvetica", size=10) # Helvetica is safer for mobile PDFs
                    pdf.cell(0, 10, txt="Sales Notes Report", ln=True, align='C')
                    pdf.ln(5)
                    for row in get_data():
                        txt = f"Date: {row[0][:10]} | Co: {row[1][:10]} | Qty: {row[3]} | Total: {row[5]}"
                        pdf.cell(0, 8, txt=txt, ln=True)
                    pdf.output(e.path)
                    show_snack("PDF Exported safely!")
                except Exception as ex:
                    show_snack(f"Export Error: {ex}")

        fp_excel = ft.FilePicker(on_result=save_excel)
        fp_pdf = ft.FilePicker(on_result=save_pdf)
        page.overlay.extend([fp_excel, fp_pdf])

        # --- Android Tab Layout ---
        tabs = ft.Tabs(
            selected_index=0,
            tabs=[
                ft.Tab(
                    text="Note",
                    icon=ft.icons.EDIT_DOCUMENT,
                    content=ft.Container(
                        padding=20,
                        content=ft.Column([
                            company_dropdown, item_dropdown, qty_input, rate_input, gem_checkbox,
                            ft.ElevatedButton("Save Note", on_click=save_note, style=ft.ButtonStyle(color="white", bgcolor="blue"))
                        ], spacing=15, scroll="auto")
                    )
                ),
                ft.Tab(
                    text="Database",
                    icon=ft.icons.STORAGE,
                    content=ft.Container(
                        padding=20,
                        content=ft.Column([
                            new_company, ft.ElevatedButton("Save Company", on_click=add_company),
                            ft.Divider(height=20, color="transparent"),
                            new_item, ft.ElevatedButton("Save Item", on_click=add_item)
                        ], spacing=10, scroll="auto")
                    )
                ),
                ft.Tab(
                    text="Export",
                    icon=ft.icons.DOWNLOAD,
                    content=ft.Container(
                        padding=20,
                        content=ft.Column([
                            ft.Text("Save reports to your phone:", weight="bold"),
                            ft.ElevatedButton("Export as Excel (.xlsx)", on_click=lambda _: fp_excel.save_file(file_name="sales.xlsx", allowed_extensions=["xlsx"]), icon=ft.icons.TABLE_CHART),
                            ft.ElevatedButton("Export as PDF (.pdf)", on_click=lambda _: fp_pdf.save_file(file_name="sales.pdf", allowed_extensions=["pdf"]), icon=ft.icons.PICTURE_AS_PDF)
                        ], spacing=20)
                    )
                )
            ]
        )

        load_dropdowns()
        page.add(tabs)
        
    except Exception as e:
        # --- FIX: THE ANTI-CRASH SCREEN ---
        # If anything breaks, it prints the error on your phone instead of going white
        page.add(
            ft.Text("App Crashed!", size=30, color="red", weight="bold"),
            ft.Text("Please take a screenshot of this error:", weight="bold"),
            ft.Text(traceback.format_exc(), size=12, selectable=True)
        )
        page.update()

ft.app(target=main)
